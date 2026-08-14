"""Analog of @lobechat/file-loaders used by official readLocalFile."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

READABLE_TEXT = frozenset(
    {
        "txt",
        "md",
        "markdown",
        "json",
        "jsonl",
        "csv",
        "tsv",
        "yaml",
        "yml",
        "xml",
        "html",
        "htm",
        "css",
        "js",
        "ts",
        "tsx",
        "jsx",
        "py",
        "rs",
        "go",
        "java",
        "c",
        "cc",
        "cpp",
        "h",
        "hpp",
        "sh",
        "bash",
        "zsh",
        "toml",
        "ini",
        "conf",
        "cfg",
        "log",
        "sql",
        "rst",
        "tex",
        "env",
        "gitignore",
        "dockerfile",
        "makefile",
        "lock",
    }
)
SPECIAL_PARSED = frozenset({"pdf", "docx", "doc", "xlsx", "xls", "pptx", "ppt", "odt", "rtf"})
SYSTEM_FILES_TO_IGNORE = frozenset({".DS_Store", "Thumbs.db", "desktop.ini", ".directory"})


def is_readable_file_type(extension: str) -> bool:
    ext = extension.lower().lstrip(".")
    if not ext:
        return True
    return ext in READABLE_TEXT or ext in SPECIAL_PARSED


def sniff_binary_file(path: str) -> tuple[bool, str]:
    data = Path(path).read_bytes()[:8192]
    if b"\x00" in data:
        return True, "contains NUL"
    return False, ""


@dataclass(slots=True)
class LoadedFile:
    content: str
    filename: str
    file_type: str
    created_time: datetime
    modified_time: datetime
    error: str = ""


def load_file(path: str) -> LoadedFile:
    p = Path(path)
    stat = p.stat()
    ext = p.suffix.lower().lstrip(".")
    created = datetime.fromtimestamp(getattr(stat, "st_ctime", stat.st_mtime))
    modified = datetime.fromtimestamp(stat.st_mtime)
    base = LoadedFile(
        content="",
        filename=p.name,
        file_type=ext or "unknown",
        created_time=created,
        modified_time=modified,
    )
    try:
        text = _load_special(p, ext)
        if text is None:
            text = p.read_text(encoding="utf-8", errors="replace")
        base.content = text
        return base
    except OSError as exc:
        base.error = str(exc)
        return base


def _load_special(path: Path, ext: str) -> str | None:
    if ext == "pdf":
        try:
            from pypdf import PdfReader

            reader = PdfReader(str(path))
            return "\n".join(page.extract_text() or "" for page in reader.pages)
        except ImportError:
            return None
    if ext == "docx":
        try:
            from docx import Document

            return "\n".join(p.text for p in Document(str(path)).paragraphs)
        except ImportError:
            return None
    if ext in {"xlsx", "xls"}:
        try:
            from openpyxl import load_workbook

            book = load_workbook(str(path), read_only=True, data_only=True)
            rows: list[str] = []
            for sheet in book.worksheets:
                rows.append(f"# {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    rows.append("\t".join("" if c is None else str(c) for c in row))
            return "\n".join(rows)
        except ImportError:
            return None
    return None
